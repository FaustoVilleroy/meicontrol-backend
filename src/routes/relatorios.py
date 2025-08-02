from flask import Blueprint, request, jsonify, g, send_file
from datetime import datetime, date
from calendar import monthrange
import os
import tempfile
from src.models.user import db
from src.models.financeiro import Receita, Despesa
from src.models.nota_fiscal import NotaFiscal
from src.models.relatorio import RelatorioMensal
from src.middleware.auth import token_required, plano_ativo_required
from src.config import Config
import json

relatorios_bp = Blueprint('relatorios', __name__)

def calcular_totais_categoria(receitas):
    """Calcula totais por categoria para MEI"""
    totais = {
        'comercio': 0,
        'servicos': 0,
        'industria': 0,
        'outros': 0
    }
    
    for receita in receitas:
        categoria = receita.categoria.lower() if receita.categoria else 'outros'
        if categoria in totais:
            totais[categoria] += float(receita.valor)
        else:
            totais['outros'] += float(receita.valor)
    
    return totais

def gerar_relatorio_mensal_pdf(dados_relatorio, mes, ano):
    """Gera PDF do relatório mensal"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        
        # Criar arquivo temporário
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        
        # Configurar documento
        doc = SimpleDocTemplate(temp_file.name, pagesize=A4)
        styles = getSampleStyleSheet()
        
        # Estilo personalizado
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Centralizado
        )
        
        story = []
        
        # Título
        story.append(Paragraph(f"RELATÓRIO MENSAL DE RECEITAS BRUTAS", title_style))
        story.append(Paragraph(f"MEI - {mes:02d}/{ano}", title_style))
        story.append(Spacer(1, 20))
        
        # Dados do MEI
        story.append(Paragraph(f"<b>Nome:</b> {dados_relatorio['usuario']['nome']}", styles['Normal']))
        story.append(Paragraph(f"<b>CNPJ:</b> {dados_relatorio['usuario']['cnpj']}", styles['Normal']))
        story.append(Paragraph(f"<b>Categoria:</b> {dados_relatorio['usuario']['categoria_mei']}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Tabela de receitas por categoria
        data = [
            ['Categoria', 'Valor (R$)'],
            ['Comércio', f"R$ {dados_relatorio['totais_categoria']['comercio']:.2f}"],
            ['Serviços', f"R$ {dados_relatorio['totais_categoria']['servicos']:.2f}"],
            ['Indústria', f"R$ {dados_relatorio['totais_categoria']['industria']:.2f}"],
            ['Outros', f"R$ {dados_relatorio['totais_categoria']['outros']:.2f}"],
            ['TOTAL GERAL', f"R$ {dados_relatorio['total_receitas']:.2f}"]
        ]
        
        table = Table(data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.beige),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        story.append(Spacer(1, 30))
        
        # Resumo de despesas
        story.append(Paragraph("<b>RESUMO DE DESPESAS</b>", styles['Heading2']))
        story.append(Paragraph(f"Total de Despesas: R$ {dados_relatorio['total_despesas']:.2f}", styles['Normal']))
        story.append(Paragraph(f"Saldo do Mês: R$ {dados_relatorio['saldo_mes']:.2f}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Observações
        story.append(Paragraph("<b>OBSERVAÇÕES IMPORTANTES:</b>", styles['Heading2']))
        story.append(Paragraph("• Este relatório deve ser entregue até o dia 20 do mês seguinte.", styles['Normal']))
        story.append(Paragraph("• Mantenha todas as notas fiscais e comprovantes arquivados por 5 anos.", styles['Normal']))
        story.append(Paragraph("• Em caso de dúvidas, consulte um contador ou o Portal do Empreendedor.", styles['Normal']))
        story.append(Spacer(1, 30))
        
        # Data de geração
        story.append(Paragraph(f"Relatório gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", styles['Normal']))
        
        # Construir PDF
        doc.build(story)
        
        return temp_file.name
        
    except Exception as e:
        print(f"Erro ao gerar PDF: {e}")
        return None

def gerar_relatorio_excel(dados_relatorio, mes, ano):
    """Gera arquivo Excel do relatório mensal"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Relatório {mes:02d}-{ano}"
        
        # Título
        ws['A1'] = f"RELATÓRIO MENSAL DE RECEITAS BRUTAS - {mes:02d}/{ano}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')
        
        # Dados do MEI
        ws['A3'] = "Nome:"
        ws['B3'] = dados_relatorio['usuario']['nome']
        ws['A4'] = "CNPJ:"
        ws['B4'] = dados_relatorio['usuario']['cnpj']
        ws['A5'] = "Categoria:"
        ws['B5'] = dados_relatorio['usuario']['categoria_mei']
        
        # Cabeçalho da tabela
        ws['A7'] = "Categoria"
        ws['B7'] = "Valor (R$)"
        
        # Estilo do cabeçalho
        for cell in ['A7', 'B7']:
            ws[cell].font = Font(bold=True)
            ws[cell].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Dados da tabela
        ws['A8'] = "Comércio"
        ws['B8'] = dados_relatorio['totais_categoria']['comercio']
        ws['A9'] = "Serviços"
        ws['B9'] = dados_relatorio['totais_categoria']['servicos']
        ws['A10'] = "Indústria"
        ws['B10'] = dados_relatorio['totais_categoria']['industria']
        ws['A11'] = "Outros"
        ws['B11'] = dados_relatorio['totais_categoria']['outros']
        ws['A12'] = "TOTAL GERAL"
        ws['B12'] = dados_relatorio['total_receitas']
        
        # Estilo do total
        ws['A12'].font = Font(bold=True)
        ws['B12'].font = Font(bold=True)
        
        # Resumo
        ws['A14'] = "Total de Despesas:"
        ws['B14'] = dados_relatorio['total_despesas']
        ws['A15'] = "Saldo do Mês:"
        ws['B15'] = dados_relatorio['saldo_mes']
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        
        # Salvar arquivo temporário
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        
        return temp_file.name
        
    except Exception as e:
        print(f"Erro ao gerar Excel: {e}")
        return None

@relatorios_bp.route('/relatorios/mensal/<int:mes>/<int:ano>', methods=['GET'])
@plano_ativo_required
def gerar_relatorio_mensal(mes, ano):
    """Gerar relatório mensal de receitas brutas"""
    try:
        # Validar mês e ano
        if mes < 1 or mes > 12:
            return jsonify({'error': 'Mês inválido'}), 400
        
        if ano < 2020 or ano > datetime.now().year:
            return jsonify({'error': 'Ano inválido'}), 400
        
        # Calcular período
        primeiro_dia = date(ano, mes, 1)
        ultimo_dia = date(ano, mes, monthrange(ano, mes)[1])
        
        # Buscar receitas do mês
        receitas = Receita.query.filter_by(user_id=g.current_user.id).filter(
            Receita.data_receita >= primeiro_dia,
            Receita.data_receita <= ultimo_dia
        ).all()
        
        # Buscar despesas do mês
        despesas = Despesa.query.filter_by(user_id=g.current_user.id).filter(
            Despesa.data_despesa >= primeiro_dia,
            Despesa.data_despesa <= ultimo_dia
        ).all()
        
        # Calcular totais
        totais_categoria = calcular_totais_categoria(receitas)
        total_receitas = sum(totais_categoria.values())
        total_despesas = sum(float(despesa.valor) for despesa in despesas)
        saldo_mes = total_receitas - total_despesas
        
        # Buscar notas fiscais do mês
        notas_fiscais = NotaFiscal.query.filter_by(user_id=g.current_user.id).filter(
            NotaFiscal.data_emissao >= primeiro_dia,
            NotaFiscal.data_emissao <= ultimo_dia
        ).all()
        
        # Preparar dados do relatório
        dados_relatorio = {
            'mes': mes,
            'ano': ano,
            'periodo': {
                'inicio': primeiro_dia.strftime('%d/%m/%Y'),
                'fim': ultimo_dia.strftime('%d/%m/%Y')
            },
            'usuario': {
                'nome': g.current_user.nome,
                'cnpj': g.current_user.cnpj,
                'categoria_mei': g.current_user.categoria_mei
            },
            'totais_categoria': totais_categoria,
            'total_receitas': total_receitas,
            'total_despesas': total_despesas,
            'saldo_mes': saldo_mes,
            'quantidade_receitas': len(receitas),
            'quantidade_despesas': len(despesas),
            'quantidade_notas_fiscais': len(notas_fiscais),
            'receitas': [receita.to_dict() for receita in receitas],
            'despesas': [despesa.to_dict() for despesa in despesas],
            'notas_fiscais': [nota.to_dict() for nota in notas_fiscais]
        }
        
        # Verificar se já existe relatório salvo
        relatorio_existente = RelatorioMensal.query.filter_by(
            user_id=g.current_user.id,
            mes=mes,
            ano=ano
        ).first()
        
        if not relatorio_existente:
            # Criar novo relatório
            novo_relatorio = RelatorioMensal(
                user_id=g.current_user.id,
                mes=mes,
                ano=ano,
                total_receitas=total_receitas,
                total_despesas=total_despesas,
                saldo_mes=saldo_mes,
                dados_json=json.dumps(dados_relatorio),
                totais_categoria_json=json.dumps(totais_categoria)
            )
            
            db.session.add(novo_relatorio)
            db.session.commit()
        else:
            # Atualizar relatório existente
            relatorio_existente.total_receitas = total_receitas
            relatorio_existente.total_despesas = total_despesas
            relatorio_existente.saldo_mes = saldo_mes
            relatorio_existente.dados_json = json.dumps(dados_relatorio)
            relatorio_existente.totais_categoria_json = json.dumps(totais_categoria)
            relatorio_existente.updated_at = datetime.utcnow()
            
            db.session.commit()
        
        return jsonify({
            'message': 'Relatório mensal gerado com sucesso',
            'relatorio': dados_relatorio
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@relatorios_bp.route('/relatorios/mensal/<int:mes>/<int:ano>/pdf', methods=['GET'])
@plano_ativo_required
def download_relatorio_mensal_pdf(mes, ano):
    """Download do relatório mensal em PDF"""
    try:
        # Buscar relatório existente
        relatorio = RelatorioMensal.query.filter_by(
            user_id=g.current_user.id,
            mes=mes,
            ano=ano
        ).first()
        
        if not relatorio:
            return jsonify({'error': 'Relatório não encontrado. Gere o relatório primeiro.'}), 404
        
        # Recuperar dados do relatório
        dados_relatorio = json.loads(relatorio.dados_json)
        
        # Gerar PDF
        pdf_path = gerar_relatorio_mensal_pdf(dados_relatorio, mes, ano)
        
        if not pdf_path:
            return jsonify({'error': 'Erro ao gerar PDF'}), 500
        
        return send_file(
            pdf_path,
            as_attachment=True,
            download_name=f'relatorio_mensal_{mes:02d}_{ano}.pdf',
            mimetype='application/pdf'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@relatorios_bp.route('/relatorios/mensal/<int:mes>/<int:ano>/excel', methods=['GET'])
@plano_ativo_required
def download_relatorio_mensal_excel(mes, ano):
    """Download do relatório mensal em Excel"""
    try:
        # Buscar relatório existente
        relatorio = RelatorioMensal.query.filter_by(
            user_id=g.current_user.id,
            mes=mes,
            ano=ano
        ).first()
        
        if not relatorio:
            return jsonify({'error': 'Relatório não encontrado. Gere o relatório primeiro.'}), 404
        
        # Recuperar dados do relatório
        dados_relatorio = json.loads(relatorio.dados_json)
        
        # Gerar Excel
        excel_path = gerar_relatorio_excel(dados_relatorio, mes, ano)
        
        if not excel_path:
            return jsonify({'error': 'Erro ao gerar Excel'}), 500
        
        return send_file(
            excel_path,
            as_attachment=True,
            download_name=f'relatorio_mensal_{mes:02d}_{ano}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@relatorios_bp.route('/relatorios/anual/<int:ano>', methods=['GET'])
@plano_ativo_required
def gerar_relatorio_anual(ano):
    """Gerar relatório anual para DASN-SIMEI"""
    try:
        # Validar ano
        if ano < 2020 or ano > datetime.now().year:
            return jsonify({'error': 'Ano inválido'}), 400
        
        # Buscar todos os relatórios mensais do ano
        relatorios_mensais = RelatorioMensal.query.filter_by(
            user_id=g.current_user.id,
            ano=ano
        ).order_by(RelatorioMensal.mes).all()
        
        # Calcular totais anuais
        total_anual_receitas = 0
        total_anual_despesas = 0
        totais_categoria_anual = {
            'comercio': 0,
            'servicos': 0,
            'industria': 0,
            'outros': 0
        }
        
        meses_com_relatorio = []
        
        for relatorio in relatorios_mensais:
            total_anual_receitas += float(relatorio.total_receitas)
            total_anual_despesas += float(relatorio.total_despesas)
            
            # Somar categorias
            totais_mes = json.loads(relatorio.totais_categoria_json)
            for categoria, valor in totais_mes.items():
                if categoria in totais_categoria_anual:
                    totais_categoria_anual[categoria] += float(valor)
            
            meses_com_relatorio.append(relatorio.mes)
        
        # Verificar se teve funcionário (campo a ser preenchido pelo usuário)
        teve_funcionario = request.args.get('teve_funcionario', 'false').lower() == 'true'
        
        # Preparar dados do relatório anual
        dados_relatorio_anual = {
            'ano': ano,
            'usuario': {
                'nome': g.current_user.nome,
                'cnpj': g.current_user.cnpj,
                'categoria_mei': g.current_user.categoria_mei
            },
            'totais_categoria_anual': totais_categoria_anual,
            'total_anual_receitas': total_anual_receitas,
            'total_anual_despesas': total_anual_despesas,
            'saldo_anual': total_anual_receitas - total_anual_despesas,
            'meses_com_relatorio': meses_com_relatorio,
            'meses_faltantes': [m for m in range(1, 13) if m not in meses_com_relatorio],
            'teve_funcionario': teve_funcionario,
            'limite_faturamento_mei': 81000.00,  # Limite MEI 2024
            'percentual_limite': (total_anual_receitas / 81000.00) * 100,
            'dentro_limite': total_anual_receitas <= 81000.00
        }
        
        return jsonify({
            'message': 'Relatório anual gerado com sucesso',
            'relatorio_anual': dados_relatorio_anual
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@relatorios_bp.route('/relatorios/historico', methods=['GET'])
@plano_ativo_required
def listar_relatorios_historico():
    """Listar histórico de relatórios mensais"""
    try:
        # Parâmetros de filtro
        page = request.args.get('page', 1, type=int)
        per_page = min(request.args.get('per_page', 12, type=int), 50)
        ano = request.args.get('ano', type=int)
        
        # Query base
        query = RelatorioMensal.query.filter_by(user_id=g.current_user.id)
        
        # Filtrar por ano se especificado
        if ano:
            query = query.filter(RelatorioMensal.ano == ano)
        
        # Ordenar por ano e mês mais recentes
        query = query.order_by(RelatorioMensal.ano.desc(), RelatorioMensal.mes.desc())
        
        # Paginação
        relatorios_paginados = query.paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        # Preparar dados para resposta
        relatorios_lista = []
        for relatorio in relatorios_paginados.items:
            relatorios_lista.append({
                'id': relatorio.id,
                'mes': relatorio.mes,
                'ano': relatorio.ano,
                'total_receitas': float(relatorio.total_receitas),
                'total_despesas': float(relatorio.total_despesas),
                'saldo_mes': float(relatorio.saldo_mes),
                'created_at': relatorio.created_at.strftime('%d/%m/%Y %H:%M'),
                'updated_at': relatorio.updated_at.strftime('%d/%m/%Y %H:%M') if relatorio.updated_at else None
            })
        
        return jsonify({
            'relatorios': relatorios_lista,
            'total': relatorios_paginados.total,
            'pages': relatorios_paginados.pages,
            'current_page': page,
            'per_page': per_page
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@relatorios_bp.route('/relatorios/status-obrigacoes', methods=['GET'])
@plano_ativo_required
def status_obrigacoes():
    """Verificar status das obrigações mensais e anuais"""
    try:
        hoje = date.today()
        ano_atual = hoje.year
        mes_atual = hoje.month
        
        # Verificar relatório do mês anterior
        if mes_atual == 1:
            mes_anterior = 12
            ano_anterior = ano_atual - 1
        else:
            mes_anterior = mes_atual - 1
            ano_anterior = ano_atual
        
        # Buscar relatório do mês anterior
        relatorio_mes_anterior = RelatorioMensal.query.filter_by(
            user_id=g.current_user.id,
            mes=mes_anterior,
            ano=ano_anterior
        ).first()
        
        # Calcular prazo do relatório mensal (dia 20)
        prazo_relatorio_mensal = date(ano_atual, mes_atual, 20)
        relatorio_mensal_em_atraso = hoje > prazo_relatorio_mensal and not relatorio_mes_anterior
        
        # Verificar DASN-SIMEI (até 31 de maio)
        prazo_dasn = date(ano_atual, 5, 31)
        dasn_em_atraso = hoje > prazo_dasn
        
        # Buscar relatórios do ano anterior para DASN
        relatorios_ano_anterior = RelatorioMensal.query.filter_by(
            user_id=g.current_user.id,
            ano=ano_anterior
        ).count()
        
        status = {
            'relatorio_mensal': {
                'mes_referencia': f"{mes_anterior:02d}/{ano_anterior}",
                'prazo': prazo_relatorio_mensal.strftime('%d/%m/%Y'),
                'entregue': bool(relatorio_mes_anterior),
                'em_atraso': relatorio_mensal_em_atraso,
                'dias_para_vencimento': (prazo_relatorio_mensal - hoje).days if hoje <= prazo_relatorio_mensal else 0
            },
            'dasn_simei': {
                'ano_referencia': ano_anterior,
                'prazo': prazo_dasn.strftime('%d/%m/%Y'),
                'relatorios_mensais_completos': relatorios_ano_anterior == 12,
                'em_atraso': dasn_em_atraso,
                'dias_para_vencimento': (prazo_dasn - hoje).days if hoje <= prazo_dasn else 0
            },
            'alertas': []
        }
        
        # Gerar alertas
        if relatorio_mensal_em_atraso:
            status['alertas'].append({
                'tipo': 'erro',
                'titulo': 'Relatório Mensal em Atraso',
                'mensagem': f'O relatório de {mes_anterior:02d}/{ano_anterior} deveria ter sido entregue até {prazo_relatorio_mensal.strftime("%d/%m/%Y")}'
            })
        elif status['relatorio_mensal']['dias_para_vencimento'] <= 5 and not relatorio_mes_anterior:
            status['alertas'].append({
                'tipo': 'aviso',
                'titulo': 'Relatório Mensal Próximo do Vencimento',
                'mensagem': f'Você tem {status["relatorio_mensal"]["dias_para_vencimento"]} dias para entregar o relatório de {mes_anterior:02d}/{ano_anterior}'
            })
        
        if dasn_em_atraso and relatorios_ano_anterior < 12:
            status['alertas'].append({
                'tipo': 'erro',
                'titulo': 'DASN-SIMEI em Atraso',
                'mensagem': f'A declaração anual de {ano_anterior} deveria ter sido entregue até {prazo_dasn.strftime("%d/%m/%Y")}'
            })
        elif status['dasn_simei']['dias_para_vencimento'] <= 30 and relatorios_ano_anterior < 12:
            status['alertas'].append({
                'tipo': 'aviso',
                'titulo': 'DASN-SIMEI Próxima do Vencimento',
                'mensagem': f'Você tem {status["dasn_simei"]["dias_para_vencimento"]} dias para entregar a declaração anual de {ano_anterior}'
            })
        
        return jsonify({'status_obrigacoes': status}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

